from twikit import Client, TooManyRequests
import asyncio
from datetime import datetime
import traceback
from random import randint
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

def processTweets(tweets, tweet_count):
    new_count = tweet_count
    tweet_data_list = []

    for tweet in tweets:
        tweet = vars(tweet)
        new_count += 1

        tweet_data = tweet.get('_data', {})
        core_data = tweet_data.get('core', {})
        user_results = core_data.get('user_results', {}).get('result', {})
        user_legacy = user_results.get('legacy', {})
        legacy_data = tweet_data.get('legacy', {})

        extracted_data = {
            'tweet_id': tweet_data.get('rest_id'),
            'created_at': legacy_data.get('created_at'),
            'full_text': legacy_data.get('full_text'),
            'conversation_id': legacy_data.get('conversation_id_str'),
            'user_id': user_results.get('rest_id'),
            'screen_name': user_legacy.get('screen_name'),
            'name': user_legacy.get('name'),
            'favorite_count': legacy_data.get('favorite_count'),
            'retweet_count': legacy_data.get('retweet_count'),
            'hashtags': legacy_data.get('entities', {}).get('hashtags', []),
            'user_mentions': legacy_data.get('entities', {}).get('user_mentions', []),
            'urls': legacy_data.get('entities', {}).get('urls', []),
            'media': legacy_data.get('entities', {}).get('media', [])
        }

        tweet_data_list.append([
            new_count, extracted_data['user_id'], extracted_data['screen_name'], extracted_data['name'],
            extracted_data['tweet_id'], extracted_data['created_at'], extracted_data['full_text'], extracted_data['retweet_count'],
            extracted_data['favorite_count'], json.dumps(extracted_data['hashtags']), 
            json.dumps(extracted_data['user_mentions']), json.dumps(extracted_data['urls']), 
            json.dumps(extracted_data['media']), extracted_data['conversation_id']
        ])

    return new_count, tweet_data_list

async def main():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tweets"
        headers = ['Index','User Id', 'Screen Name', 'Name', 'Tweet Id','Created At', 'Text', 'Retweet Count', 'Likes', 'Hashtags', 'User Mentions', 'Urls', 'Media', 'Conversation Id']
        ws.append(headers)
        
        client = Client('en-US')
        client.load_cookies('cookies.json') 
        
        tweet_count = 0
        tweets = None
        QUERY = "Python"
        
        while tweet_count < 50:
            try:
                if tweets is None:
                    print(f"{datetime.now()} - getting Tweets...")
                    tweets = await client.search_tweet(QUERY, product='Latest')
                else:
                    print(f"{datetime.now()} - getting next batch of Tweets...")
                    tweets = await tweets.next()
            except TooManyRequests as e:
                rate_limit_reset = datetime.fromtimestamp(e.rate_limit_reset)
                wait_time = (rate_limit_reset - datetime.now()).total_seconds() + 20
                print(f"{datetime.now()} - Rate limit reached. Waiting for {wait_time} seconds.")
                await asyncio.sleep(wait_time)
                continue
            
            if not tweets:
                print(f"{datetime.now()} No more Tweets found")

            tweet_count, tweet_data_list = processTweets(tweets, tweet_count)
            
            for tweet_data in tweet_data_list:
                ws.append(tweet_data)
            
            wait = randint(10, 20)
            print(f"Waiting for {wait} seconds before the next batch...")
            await asyncio.sleep(wait)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        wb.save("Python.xlsx")

    except Exception as e:
        print(f"An error occurred: {e}")
        traceback.print_exc()
    finally:
        print(f"{datetime.now()} finished execution total tweets. tweet count {tweet_count}")

if __name__ == "__main__":
    asyncio.run(main())
